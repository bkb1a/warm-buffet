"""Parse the real broker transaction exports (Binck + Bolero) and upsert them
into Supabase `transactions`. Also importable: parse_all() returns the parsed
list so other scripts (export_dashboard) can use it without the DB.

Binck  (Documenten/transacties-binck.xlsx): qty + name in Omschrijving.
Bolero (Documenten/Transacties Bolero ...): no quantities; qty is derived from
the first meeting snapshot after the purchase when the amounts match.

Bolero date quirk: dd-mm-yyyy cells that Excel could read as mm-dd were stored
swapped (e.g. "11-03-2026" = 11 March -> stored as 3 Nov). Datetime cells are
therefore un-swapped; string cells are parsed as dd-mm-yyyy.
"""
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

from common import ROOT, Supa
from ingest_rendement import NAME_MAP

DOCS = ROOT / "Documenten"


def bolero_file():
    """Newest Bolero export in Documenten/ (any 'Transacties Bolero*.xlsx')."""
    files = sorted(DOCS.glob("Transacties Bolero*.xlsx"), key=lambda p: p.stat().st_mtime)
    if not files:
        raise FileNotFoundError("Geen 'Transacties Bolero*.xlsx' in Documenten/")
    return files[-1]
STRIP = re.compile(r"\s*-(Stock split|Acquired|right|[Dd]elisted)-\s*$")

# Corrections & additions confirmed against the Bolero app order history
# (screenshots 2026-08-07). Overrides win over derived share counts.
SHARE_OVERRIDES = {("2026-01-13", "BARRICK MINING CORPORATION", "sell"): 77}
# Orders newer than the xlsx export; amount estimated from the week price.
# Replace by the real row (and drop these) at the next Bolero xlsx export.
EXTRA_TXNS = [{"txn_date": "2026-08-07", "side": "buy",
               "holding_name": "ISHAR.III PLC CORE MSCI WORLD KAP",
               "raw_name": "ISHAR.III PLC CORE MSCI WORLD (app, geschat bedrag)",
               "shares": 19, "amount_eur": None, "price_eur": None, "source": "bolero"}]

EXTRA_ALIASES = {
    "Exmar NV": "Exmar NV.",
    "iShares Global Water UCITS ETF": "ISHARES II PLC S&P GLOB. WATER FD D",
}


def norm(s):
    return re.sub(r"[^a-z0-9]", "", s.lower())


def canonical(raw, canon_names):
    name = STRIP.sub("", raw).strip()
    name = EXTRA_ALIASES.get(name, NAME_MAP.get(name, name))
    if name in canon_names:
        return name
    n = norm(name)
    for c in canon_names:  # broker exports truncate names; match on normalized prefix
        if norm(c).startswith(n) or n.startswith(norm(c)):
            return c
    return name  # pre-club items (Delhaize, Asian Citrus, ...) keep their raw name


def parse_binck(canon_names):
    ws = openpyxl.load_workbook(DOCS / "transacties-binck.xlsx", data_only=True).active
    out = []
    for row in list(ws.iter_rows(values_only=True))[1:]:
        _, d, _, ttype, desc, mut, _ = row
        if ttype not in ("Aankoop", "Verkoop") or not desc:
            continue
        m = re.match(r"^([\d.]+)\s+(.+)$", str(desc).strip())
        if not m:
            continue
        qty = float(m.group(1).replace(".", ""))
        amount = abs(float(mut))
        out.append({"txn_date": str(d.date() if isinstance(d, datetime) else d),
                    "side": "buy" if ttype == "Aankoop" else "sell",
                    "holding_name": canonical(m.group(2), canon_names),
                    "raw_name": m.group(2).strip(), "shares": qty,
                    "amount_eur": amount,
                    "price_eur": round(amount / qty, 4) if qty else None,
                    "source": "binck"})
    return out


def bolero_date(v):
    if isinstance(v, datetime):
        return date(v.year, v.day, v.month).isoformat()  # un-swap dd/mm
    d, m, y = str(v).strip().split("-")
    return f"{y}-{m}-{d}"


def parse_bolero(canon_names):
    ws = openpyxl.load_workbook(bolero_file(), data_only=True).active
    out = []
    for row in list(ws.iter_rows(values_only=True))[1:]:
        d, ttype, kind, details, value, _ = row
        if kind in (None, "Cash") or not details:
            continue
        side = "buy" if str(ttype).startswith("Aankoop") else "sell"
        raw = re.sub(r"^(Aankoop|Verkoop) Online\s*", "", str(details))
        raw = re.sub(r"\s*\(\w+\)?\s*$", "", raw).strip()
        amount = abs(float(str(value).replace(".", "").replace(",", ".").replace("+", "")))
        out.append({"txn_date": bolero_date(d), "side": side,
                    "holding_name": canonical(raw, canon_names), "raw_name": raw,
                    "shares": None, "amount_eur": amount, "price_eur": None,
                    "source": "bolero"})
    return out


def derive_bolero_shares(txns, snapshots, supa=None):
    """Fill Bolero share counts: first from the meeting snapshot whose purchase
    value matches the transaction amount (within 2%); otherwise estimate from
    the prices_weekly week price (rounded to whole shares — fees make the
    division slightly off)."""
    for t in txns:
        if t["source"] != "bolero" or t["shares"]:
            continue
        cands = [s for s in snapshots
                 if s["holding_name"] == t["holding_name"] and s["meeting_date"] >= t["txn_date"]
                 and s.get("shares")]
        for s in sorted(cands, key=lambda s: s["meeting_date"]):
            ref = s.get("purchase_value_eur") or s.get("value_eur")
            if ref and abs(t["amount_eur"] / ref - 1) < 0.02:
                t["shares"] = s["shares"]
                t["price_eur"] = round(t["amount_eur"] / s["shares"], 4)
                break

    missing = [t for t in txns if t["source"] == "bolero" and not t["shares"]]
    if not (missing and supa):
        return
    tickers = {h["name"]: h["ticker"] for h in supa.select("holdings", {"select": "name,ticker"})}
    for t in missing:
        tk = tickers.get(t["holding_name"])
        if not tk:
            continue
        from datetime import date, timedelta
        d = date.fromisoformat(t["txn_date"])
        monday = (d - timedelta(days=d.weekday())).isoformat()
        rows = supa.select("prices_weekly", {"select": "avg_price",
                                             "ticker": f"eq.{tk}", "week_start": f"eq.{monday}"})
        if rows:
            px = float(rows[0]["avg_price"])
            est = round(t["amount_eur"] / px)
            # snap to the sheet's share count when the estimate is within 3%
            # (fees skew the plain division, e.g. IWDA: 90 estimated vs 89 real)
            snap = next((s2["shares"] for s2 in sorted(
                (s2 for s2 in snapshots
                 if s2["holding_name"] == t["holding_name"]
                 and s2["meeting_date"] >= t["txn_date"] and s2.get("shares")),
                key=lambda s2: s2["meeting_date"])), None)
            if est and snap and abs(snap / est - 1) <= 0.03:
                est = snap
            if est:
                t["shares"] = est
                t["price_eur"] = round(t["amount_eur"] / est, 4)


def parse_all(snapshots=None):
    s = Supa()
    canon = {h["name"] for h in s.select("holdings", {"select": "name"})}
    snapshots = snapshots or s.select("snapshots", {"select": "holding_name,meeting_date,shares,purchase_value_eur,value_eur",
                                                    "order": "meeting_date.asc", "limit": "10000"})
    txns = parse_binck(canon) + parse_bolero(canon) + [dict(t) for t in EXTRA_TXNS]
    for t in txns:
        ov = SHARE_OVERRIDES.get((t["txn_date"], t["holding_name"], t["side"]))
        if ov:
            t["shares"] = ov
            t["price_eur"] = round(t["amount_eur"] / ov, 4) if t["amount_eur"] else None
    for t in txns:
        if t["amount_eur"] is None and t["shares"]:  # estimate from the week price
            tk = {h["name"]: h["ticker"] for h in s.select("holdings", {"select": "name,ticker"})}.get(t["holding_name"])
            from datetime import date, timedelta
            d = date.fromisoformat(t["txn_date"])
            monday = (d - timedelta(days=d.weekday())).isoformat()
            rows = s.select("prices_weekly", {"select": "avg_price", "ticker": f"eq.{tk}",
                                              "week_start": f"eq.{monday}"})
            if rows:
                t["price_eur"] = round(float(rows[0]["avg_price"]), 4)
                t["amount_eur"] = round(t["price_eur"] * t["shares"], 2)
    derive_bolero_shares(txns, snapshots, supa=s)
    return sorted([t for t in txns if t["amount_eur"] is not None], key=lambda t: t["txn_date"])


def main():
    txns = parse_all()
    unmatched = sorted({t["holding_name"] for t in txns} -
                       {h["name"] for h in Supa().select("holdings", {"select": "name"})})
    print(f"{len(txns)} transactions ({sum(1 for t in txns if t['source']=='binck')} binck, "
          f"{sum(1 for t in txns if t['source']=='bolero')} bolero)")
    if unmatched:
        print("not linked to a holding (pre-club era):", unmatched)
    if "--dry-run" in sys.argv:
        for t in txns:
            print(f"  {t['txn_date']} {t['side']:4} {t['holding_name'][:34]:36} "
                  f"{t['shares'] or '?':>7} x {t['price_eur'] or '?':<9} = {t['amount_eur']:.2f} ({t['source']})")
        return
    Supa().upsert("transactions", txns, on_conflict="source,txn_date,raw_name,side,amount_eur")
    print("Upserted into Supabase.")


if __name__ == "__main__":
    main()
