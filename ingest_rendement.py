"""Ingest 't Warm Buffet Rendement.xlsx into Supabase.

Each dd-mm-YYYY sheet is one meeting snapshot: per-holding rows, portfolio
totals, currency exposure, and benchmark returns. Idempotent: upserts on
(meeting_date, holding_name) / meeting_date, so re-runs never duplicate.

Usage:
    python ingest_rendement.py            # ingest all meeting sheets
    python ingest_rendement.py --dry-run  # parse and print, no writes
"""
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from common import ROOT, Supa

XLSX = ROOT / "Documenten" / "t Warm Buffet Rendement.xlsx"
DATE_RE = re.compile(r"^\d{2}-\d{2}-\d{4}$")

# Holdings were listed under different names across sheet eras (broker rename,
# corporate renames, sold-markers). Canonical name = the most recent spelling.
NAME_MAP = {
    "Ageas": "AGEAS NV/SA",
    "Arcadis": "ARCADIS NV",
    "Barrick Gold": "BARRICK MINING CORPORATION",
    "Barrick Gold Corp": "BARRICK MINING CORPORATION",
    "Berkshire Hathaway Inc -B-": "BERKSHIRE HATHAWAY INC. -B-",
    "Brederode SA": "BREDERODE",
    "Flow Traders": "FLOW TRADERS LTD",
    "Henkel AG & Co. KGaA - VERKOCHT": "Henkel AG & Co. KGaA",
    "Hornbach Baumarkt AG..": "Hornbach Baumarkt AG",
    "Montea SCA": "MONTEA NV GVV",
    "Ter Beke": "What's Cooking (Ter Beke)",
    "Vanguard EM ETF": "Vanguard FTSE Emerging Markets ETF",
    "Vanguard FTSE Emerging Markets -ETF-": "Vanguard FTSE Emerging Markets ETF",
    "Volkswagen": "Volkswagen AG",
    "iShares Core MSCI World UCITS ETF": "ISHAR.III PLC CORE MSCI WORLD KAP",
    "iShares S&P Global Water ETF": "ISHARES II PLC S&P GLOB. WATER FD D",
    "Xtrackers Physical Silver ETC (EUR)": "XTRACKER SILVER ETC EUR ETC",
}

# The sheet layout evolved over the years, so columns are resolved from the
# header row ("Effect", "Positie", ...) instead of fixed positions.
HEADER_FIELDS = [  # (field, matcher on lowercased header) — first match wins
    ("shares", lambda h: h == "positie"),
    ("price", lambda h: h == "koers"),
    ("prev_price", lambda h: h.startswith("prev koers") or h == "prev"),
    ("value_eur", lambda h: h.startswith("val (")),
    ("weight", lambda h: h == "gewicht"),
    ("return_since_prev", lambda h: h.startswith("rendement")),
    ("purchase_price", lambda h: h.startswith("aank.koers")),
    ("purchase_value_eur", lambda h: h.startswith("aank.waarde")),
    ("unrealized_pnl_pct", lambda h: h.startswith(("onger. rend. %", "ger. rend. %"))),
    ("unrealized_pnl_eur", lambda h: h.startswith(("onger. rend.", "ger. rend."))),
]


def num(v):
    if not isinstance(v, (int, float)) or isinstance(v, bool):
        return None
    return round(float(v), 6)


def cell(row, i):
    return row[i] if i < len(row) else None


def label_of(row):
    v = cell(row, 0)
    return str(v).strip().lower() if isinstance(v, str) else ""


def parse_sheet(ws, meeting_date):
    rows = list(ws.iter_rows(values_only=True))
    header_i = next((i for i, r in enumerate(rows) if label_of(r) == "effect"), None)
    if header_i is None:
        return None

    header = [str(c).strip().lower() if isinstance(c, str) else "" for c in rows[header_i]]
    cols = {}
    for field, match in HEADER_FIELDS:
        cols[field] = next((i for i, h in enumerate(header)
                            if h and match(h) and i not in cols.values()), None)
    if cols["value_eur"] is None:  # older sheets: plain "Val" column holds EUR value
        cols["value_eur"] = next((i for i, h in enumerate(header) if h == "val"), None)
    val_col = cols["value_eur"]

    snapshots, totals = [], {"meeting_date": meeting_date.isoformat()}
    NON_HOLDING = ("totaal", "noot", "cash", "eur", "usd", "s&p", "msci",
                   "eurostoxx", "totale", "blootstelling", "rendementen",
                   "vermogen", "verkochte", "effectief", "saxo", "kbc", "val ")
    for row in rows[header_i + 1:]:
        label = label_of(row)
        shares = num(cell(row, cols["shares"])) if cols["shares"] is not None else None
        if label and shares is not None and not label.startswith(NON_HOLDING):
            ccy = cell(row, 1)
            raw_name = str(cell(row, 0)).strip()
            snap = {"meeting_date": meeting_date.isoformat(),
                    "holding_name": NAME_MAP.get(raw_name, raw_name),
                    "currency": ccy.strip() if isinstance(ccy, str) and len(ccy.strip()) == 3 else None,
                    "shares": shares}
            for k, i in cols.items():
                if k != "shares":
                    snap[k] = num(cell(row, i)) if i is not None else None
            snapshots.append(snap)
        elif label.startswith("totaal effecten"):
            totals["securities_value_eur"] = num(cell(row, val_col))
        elif label == "cash":
            totals["cash_eur"] = num(cell(row, val_col))
        elif label.startswith("totaal portefeuille") or label == "totaal":
            totals.setdefault("total_value_eur",
                              num(cell(row, val_col)) or num(cell(row, 2)))
        elif label == "eur":
            totals["eur_exposure"] = num(cell(row, 3))
        elif label == "usd":
            totals["usd_exposure"] = num(cell(row, 3))
        elif label.startswith("totale portefeuille"):
            totals["return_portfolio"] = num(cell(row, 2))
        elif label == "s&p 500":
            totals["return_sp500"] = num(cell(row, 2))
        elif label == "msci world":
            totals["return_msci_world"] = num(cell(row, 2))
        elif label.startswith("eurostoxx"):
            totals["return_eurostoxx50"] = num(cell(row, 2))
    return merge_lots(snapshots), totals


def merge_lots(snapshots):
    """A holding bought in tranches appears as multiple rows on one sheet
    (e.g. Vanguard FTSE EM 2019-2023) — combine them into one position."""
    out = {}
    for s in snapshots:
        prev = out.get(s["holding_name"])
        if not prev:
            out[s["holding_name"]] = s
            continue
        for k in ("shares", "value_eur", "weight", "purchase_value_eur", "unrealized_pnl_eur"):
            if s[k] is not None:
                prev[k] = (prev[k] or 0) + s[k]
        if prev["purchase_value_eur"] and prev["shares"]:
            prev["purchase_price"] = round(prev["purchase_value_eur"] / prev["shares"], 6)
        if prev["purchase_value_eur"] and prev["unrealized_pnl_eur"] is not None:
            prev["unrealized_pnl_pct"] = round(prev["unrealized_pnl_eur"] / prev["purchase_value_eur"], 6)
    return list(out.values())


def main():
    dry = "--dry-run" in sys.argv
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    meeting_sheets = [(datetime.strptime(n, "%d-%m-%Y").date(), n)
                      for n in wb.sheetnames if DATE_RE.match(n)]
    meeting_sheets.sort()

    all_snaps, all_totals = [], []
    for d, name in meeting_sheets:
        parsed = parse_sheet(wb[name], d)
        if not parsed:
            print(f"  ! {name}: no 'Effect' header found, skipped")
            continue
        snaps, totals = parsed
        all_snaps.extend(snaps)
        all_totals.append(totals)
        print(f"  {name}: {len(snaps)} holdings, total €{totals.get('total_value_eur')}")

    latest_date = meeting_sheets[-1][0].isoformat()
    latest_names = {s["holding_name"] for s in all_snaps if s["meeting_date"] == latest_date}
    holdings = [{"name": n, "currency": next((s["currency"] for s in all_snaps
                                              if s["holding_name"] == n and s["currency"]), "EUR"),
                 "active": n in latest_names}
                for n in sorted({s["holding_name"] for s in all_snaps})]

    print(f"\n{len(all_totals)} meetings, {len(all_snaps)} snapshot rows, "
          f"{len(holdings)} distinct holdings ({len(latest_names)} active)")
    if dry:
        return

    supa = Supa()
    supa.upsert("holdings", holdings, on_conflict="name")
    supa.upsert("snapshots", all_snaps, on_conflict="meeting_date,holding_name")
    keys = {k for t in all_totals for k in t}  # PostgREST bulk rows need uniform keys
    supa.upsert("portfolio_totals", [{k: t.get(k) for k in keys} for t in all_totals],
                on_conflict="meeting_date")
    print("Ingested into Supabase.")


if __name__ == "__main__":
    main()
